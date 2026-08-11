"""用統一編號補公司登記資料。

這一步的價值不在資本額，在「這家公司還在不在」。名錄網站不會把倒掉的會員
刪掉，所以一份剛爬回來的名單裡混著早就解散、撤銷、廢止的公司。

測試不碰網路：對方的回應形狀已經固定，用假的 fetcher 餵真實的回應內容就
測得到全部的分支——包含最重要的那個，「回來的不是 JSON 而是一頁忙碌中的
HTML」。那不是假想的情況，實測時遇過。
"""

from __future__ import annotations

from datetime import date, timedelta

import pytest

from core.legal import OPEN_DATA_ATTRIBUTION
from crawler.fetcher import FetchResult
from crawler.registry import (
    RECHECK_AFTER_DAYS,
    RegistryBusy,
    best_name_match,
    enrich_registrations,
    lookup,
    lookup_by_name,
    name_query_url,
    parse_name_response,
    parse_response,
    query_url,
    roc_to_date,
)
from database.models import Company, now
from database.repository import CompanyRepository

#: 台積電。實際打過這支 API 拿回來的內容，欄位名稱與型別都照原樣。
TSMC_JSON = """[{"Business_Accounting_NO":"22099131","Company_Status_Desc":"核准設立",
"Company_Name":"台灣積體電路製造股份有限公司","Capital_Stock_Amount":280500000000,
"Paid_In_Capital_Amount":259323700670,"Responsible_Name":"魏哲家",
"Company_Location":"新竹科學園區新竹市力行六路8號","Company_Setup_Date":"0760221",
"Register_Organization_Desc":"國家科學及技術委員會新竹科學園區管理局"}]"""

#: 一家已經撤銷的公司。這種才是這個功能真正要抓出來的。
REVOKED_JSON = """[{"Business_Accounting_NO":"04595257","Company_Status_Desc":"撤銷",
"Company_Name":"擎天金屬股份有限公司","Capital_Stock_Amount":150000000,
"Company_Location":"桃園市桃園區民生北路１６２號","Company_Setup_Date":"0650923"}]"""

#: 尖峰時間回的東西。HTTP 是 200，內容是網頁。
BUSY_HTML = "<html><body><h1>系統忙碌中，請稍後再試</h1></body></html>"


class _Fetcher:
    """照順序吐出預先準備好的回應，並記下被要求了哪些網址。"""

    def __init__(self, *bodies: str) -> None:
        self._bodies = list(bodies)
        self.urls: list[str] = []
        self.closed = False

    def fetch(self, url: str, **_kwargs) -> FetchResult:
        self.urls.append(url)
        body = self._bodies.pop(0) if self._bodies else ""
        return FetchResult(url=url, status_code=200, html=body)

    def close(self) -> None:
        self.closed = True


# ------------------------------------------------------------ 民國日期


@pytest.mark.parametrize(
    "roc, expected",
    [("0760221", date(1987, 2, 21)), ("1150618", date(2026, 6, 18)),
     ("0650923", date(1976, 9, 23))],
)
def test_roc_dates_are_converted(roc, expected):
    assert roc_to_date(roc) == expected


@pytest.mark.parametrize("bad", ["", None, "abc", "76221", "0769999"])
def test_an_unreadable_date_is_dropped_not_raised(bad):
    """設立日期是順帶的資訊。為了一格怪日期讓整批補完停下來不值得。"""
    assert roc_to_date(bad) is None


# ------------------------------------------------------------ 解析回應


def test_a_real_response_is_parsed():
    registration = parse_response(TSMC_JSON, "22099131")

    assert registration is not None
    assert registration.company_name == "台灣積體電路製造股份有限公司"
    assert registration.status == "核准設立"
    assert registration.is_active
    assert registration.capital_amount == 280_500_000_000
    assert registration.paid_in_capital == 259_323_700_670
    assert registration.responsible_name == "魏哲家"
    assert registration.setup_date == date(1987, 2, 21)


def test_a_revoked_company_is_not_active():
    registration = parse_response(REVOKED_JSON, "04595257")

    assert registration is not None
    assert registration.status == "撤銷"
    assert not registration.is_active


def test_an_empty_body_means_no_such_company():
    """查無此統編時對方回 200 加空白內容，那不是錯誤。"""
    assert parse_response("", "12345675") is None
    assert parse_response("[]", "12345675") is None


def test_a_busy_page_is_reported_as_busy_not_as_no_data():
    """把忙線當成「查無此公司」會留下一個永遠不會再查的錯誤結論。"""
    with pytest.raises(RegistryBusy):
        parse_response(BUSY_HTML, "22099131")


def test_the_query_url_encodes_the_filter():
    url = query_url("22099131")

    assert "22099131" in url
    assert " " not in url  # OData 的 filter 含空白，沒編碼會被伺服器拒絕


# --------------------------------------------------------- 用名稱查詢
#
# 統編查詢用的那個資料集不支援名稱查詢（實測回空陣列）。名稱查詢走的是
# 另一個資料集，而它有一個不寫在文件裡、少了就一定回空陣列的條件。

#: 名稱查詢的回應。比統編查詢多一個 ``Company_Status`` 代碼欄位。
TSMC_BY_NAME_JSON = """[{"Business_Accounting_NO":"22099131",
"Company_Name":"台灣積體電路製造股份有限公司","Company_Status":"01",
"Company_Status_Desc":"核准設立","Capital_Stock_Amount":280500000000,
"Responsible_Name":"魏哲家","Company_Location":"新竹科學園區新竹市力行六路8號",
"Company_Setup_Date":"0760221"}]"""

#: 模糊查詢「台積電」實際會回來的東西——沒有一家是台積電。
NEAR_MISSES_JSON = """[
{"Business_Accounting_NO":"54900838","Company_Name":"台積電機有限公司",
 "Company_Status":"01","Company_Status_Desc":"核准設立","Responsible_Name":"王志聰"},
{"Business_Accounting_NO":"90312187","Company_Name":"台積電梯有限公司",
 "Company_Status":"01","Company_Status_Desc":"核准設立","Responsible_Name":"余湧筠"}
]"""


def test_the_name_query_url_carries_the_status_condition():
    """``and Company_Status eq 01`` 不能拿掉。

    少了它對方回的是空陣列，而不是錯誤——所以看起來會像「這家公司沒登記」，
    而不是「查詢寫錯了」。實測過：帶條件回 1 筆，不帶條件回 0 筆。
    """
    url = name_query_url("台灣積體電路製造")

    assert "Company_Status" in url or "Company_Status".replace("_", "%5F") in url
    assert "eq%2001" in url or "eq+01" in url or "eq%2001" in url
    assert " " not in url


def test_the_name_query_uses_the_other_dataset():
    """兩個資料集的用途不同，指錯了會安靜地一筆都查不到。"""
    assert name_query_url("甲") != query_url("22099131")
    assert "6BBA2268" in name_query_url("甲")


def test_a_name_response_is_parsed_into_a_list():
    results = parse_name_response(NEAR_MISSES_JSON)

    assert [r.company_name for r in results] == ["台積電機有限公司", "台積電梯有限公司"]
    assert results[0].tax_id == "54900838"


def test_an_empty_name_response_is_not_an_error():
    assert parse_name_response("") == []
    assert parse_name_response("[]") == []


def test_a_busy_response_is_still_detected_on_the_name_path():
    with pytest.raises(RegistryBusy):
        parse_name_response(BUSY_HTML)


def test_the_legal_suffix_does_not_have_to_match():
    """名錄上的簡稱跟登記全名常常對不起來。"""
    candidates = parse_name_response(TSMC_BY_NAME_JSON)

    matched = best_name_match("台灣積體電路製造", candidates)

    assert matched is not None
    assert matched.tax_id == "22099131"


def test_a_near_miss_is_not_a_match():
    """把統編補到錯的公司上，比留白糟得多——留白至少看得出來是缺的。"""
    assert best_name_match("台積電", parse_name_response(NEAR_MISSES_JSON)) is None


def test_an_empty_name_matches_nothing():
    assert best_name_match("", parse_name_response(TSMC_BY_NAME_JSON)) is None


def test_a_name_lookup_goes_out_and_comes_back():
    fetcher = _Fetcher(TSMC_BY_NAME_JSON)

    registration = lookup_by_name("台灣積體電路製造股份有限公司", fetcher)

    assert registration is not None
    assert registration.tax_id == "22099131"
    assert registration.responsible_name == "魏哲家"
    assert registration.is_active


def test_a_name_lookup_that_only_finds_near_misses_gives_nothing():
    fetcher = _Fetcher(NEAR_MISSES_JSON)

    assert lookup_by_name("台積電", fetcher) is None


def test_a_blank_name_never_sends_a_request():
    fetcher = _Fetcher(TSMC_BY_NAME_JSON)

    assert lookup_by_name("   ", fetcher) is None
    assert fetcher.urls == []


# ------------------------------------------------------------ 單筆查詢


def test_a_lookup_goes_out_and_comes_back():
    fetcher = _Fetcher(TSMC_JSON)

    registration = lookup("22099131", fetcher)

    assert registration is not None
    assert len(fetcher.urls) == 1


def test_a_tax_id_that_fails_its_checksum_is_never_sent():
    """統編有檢查碼。過不了的一定查不到，不值得為它送一次請求。"""
    fetcher = _Fetcher(TSMC_JSON)

    assert lookup("12345678", fetcher) is None
    assert fetcher.urls == []


# ------------------------------------------------------------ 整批補完


def _add(session, **fields) -> Company:
    name = fields.pop("company_name", "某某企業有限公司")
    # dedupe_key 有唯一索引，留白的話第二筆就進不去了。
    company = Company(company_name=name, dedupe_key=f"name:{name}", **fields)
    session.add(company)
    session.commit()
    return company


def test_a_company_is_filled_in_from_its_registration(db_session):
    company = _add(db_session, tax_id="22099131")
    fetcher = _Fetcher(TSMC_JSON)

    summary = enrich_registrations(fetcher=fetcher)

    db_session.refresh(company)
    assert summary.matched == 1 and summary.updated == 1
    assert company.registration_status == "核准設立"
    assert company.capital_amount == 280_500_000_000
    assert company.contact_person == "魏哲家"
    assert company.address == "新竹科學園區新竹市力行六路8號"
    assert company.extra_fields["登記名稱"] == "台灣積體電路製造股份有限公司"
    assert company.registration_checked_at is not None


def test_a_defunct_company_is_counted_so_the_user_hears_about_it(db_session):
    _add(db_session, company_name="擎天金屬股份有限公司", tax_id="04595257")

    summary = enrich_registrations(fetcher=_Fetcher(REVOKED_JSON))

    assert summary.defunct == 1


def test_existing_details_are_not_overwritten(db_session):
    """官方登記的負責人不一定是使用者談的窗口。使用者填過的東西不能被蓋掉。"""
    company = _add(
        db_session, tax_id="22099131", contact_person="陳採購", address="使用者填的地址"
    )

    enrich_registrations(fetcher=_Fetcher(TSMC_JSON))

    db_session.refresh(company)
    assert company.contact_person == "陳採購"
    assert company.address == "使用者填的地址"


def test_the_registration_status_is_always_refreshed(db_session):
    """反過來，登記狀態與資本額本來就是這一步負責維護的。舊值一定比較不準。"""
    company = _add(db_session, tax_id="04595257", registration_status="核准設立")

    enrich_registrations(fetcher=_Fetcher(REVOKED_JSON))

    db_session.refresh(company)
    assert company.registration_status == "撤銷"


def test_companies_without_a_tax_id_are_skipped_not_guessed(db_session):
    """這支 API 只能用統編查。沒有統編就是查不了，不能拿名稱去猜。"""
    _add(db_session, company_name="沒有統編的公司")
    fetcher = _Fetcher(TSMC_JSON)

    summary = enrich_registrations(fetcher=fetcher)

    assert summary.skipped_no_tax_id == 1
    assert fetcher.urls == []


def test_a_company_already_checked_is_not_asked_again(db_session):
    company = _add(db_session, tax_id="22099131")
    company.registration_checked_at = now()
    db_session.commit()
    fetcher = _Fetcher(TSMC_JSON)

    summary = enrich_registrations(fetcher=fetcher)

    assert summary.considered == 0
    assert fetcher.urls == []


def test_a_stale_check_is_refreshed(db_session):
    """公司登記不常變，但「解散」一旦發生就很重要。放太久要重查一次。"""
    company = _add(db_session, tax_id="22099131")
    company.registration_checked_at = now() - timedelta(days=RECHECK_AFTER_DAYS + 1)
    db_session.commit()

    summary = enrich_registrations(fetcher=_Fetcher(TSMC_JSON))

    assert summary.considered == 1


def test_a_company_with_no_registration_is_marked_as_checked(db_session):
    """查無此統編也算查過了，否則每一次補完都會再問一遍同一批。"""
    company = _add(db_session, tax_id="12345675")

    summary = enrich_registrations(fetcher=_Fetcher(""))

    db_session.refresh(company)
    assert summary.not_found == 1
    assert company.registration_checked_at is not None


def test_a_busy_reply_leaves_the_company_unchecked_so_it_is_retried(db_session):
    """忙線不是答案。標記成已查過的話，這一筆就再也不會被問了。"""
    company = _add(db_session, tax_id="22099131")

    summary = enrich_registrations(fetcher=_Fetcher(BUSY_HTML))

    db_session.refresh(company)
    assert summary.busy == 1
    assert company.registration_checked_at is None


def test_one_bad_company_does_not_stop_the_batch(db_session):
    """一趟補完可能跑幾百家。第三家壞掉不該讓後面的都不做。"""
    _add(db_session, company_name="甲", tax_id="22099131")
    _add(db_session, company_name="乙", tax_id="04595257")

    summary = enrich_registrations(fetcher=_Fetcher(BUSY_HTML, REVOKED_JSON))

    assert summary.busy == 1
    assert summary.matched == 1


def test_a_borrowed_fetcher_is_not_closed(db_session):
    """呼叫端自己開的 fetcher 由呼叫端關。這裡關掉會害下一段程式沒得用。"""
    _add(db_session, tax_id="22099131")
    fetcher = _Fetcher(TSMC_JSON)

    enrich_registrations(fetcher=fetcher)

    assert not fetcher.closed


def test_cancelling_stops_the_batch(db_session):
    import threading

    _add(db_session, company_name="甲", tax_id="22099131")
    _add(db_session, company_name="乙", tax_id="04595257")
    cancel = threading.Event()
    cancel.set()

    summary = enrich_registrations(fetcher=_Fetcher(TSMC_JSON), cancel_event=cancel)

    assert summary.matched == 0


# ------------------------------------------------------------ 待處理家數


def test_the_pending_count_only_counts_what_can_actually_be_looked_up(db_session):
    _add(db_session, company_name="有統編", tax_id="22099131")
    _add(db_session, company_name="沒統編")

    assert CompanyRepository(db_session).count_registrable() == 1


# ------------------------------------------------------------ 顯名標示


def test_the_attribution_names_the_source_agency():
    """條款規定未盡顯名標示義務者視為自始未取得授權。這一行不能是空的，
    也不能沒有提到資料提供機關。"""
    assert "經濟部商業司" in OPEN_DATA_ATTRIBUTION
    assert "政府資料開放授權條款" in OPEN_DATA_ATTRIBUTION


def _view(**fields):
    from core.schemas import CompanyView

    base = dict(id=1, company_name="某某企業有限公司")
    base.update(fields)
    return CompanyView(**base)


def test_data_without_any_registration_needs_no_attribution():
    """只匯出公司名稱與信箱時硬掛一行來源聲明只是雜訊。"""
    from exporter.base import registry_attribution

    assert registry_attribution([_view(email="a@b.com")]) is None


def test_one_enriched_row_makes_the_whole_file_need_the_attribution():
    from exporter.base import registry_attribution

    rows = [_view(id=1), _view(id=2, registration_status="核准設立")]

    assert registry_attribution(rows) == OPEN_DATA_ATTRIBUTION


@pytest.mark.parametrize("fmt", ["csv", "json", "excel"])
def test_every_export_format_carries_the_attribution(fmt, tmp_path, patch_config):
    """三種格式都要帶著它。少一種就是那一種格式的檔案從一開始就沒有授權。"""
    from exporter.service import get_exporter

    rows = [_view(registration_status="核准設立", capital_amount=280_500_000_000)]

    target = get_exporter(fmt).export(rows, tmp_path / f"out.{fmt}")

    if fmt == "excel":
        import openpyxl

        book = openpyxl.load_workbook(target)
        assert "資料來源" in book.sheetnames
        assert OPEN_DATA_ATTRIBUTION in str(book["資料來源"]["A2"].value)
    else:
        assert OPEN_DATA_ATTRIBUTION in target.read_text(encoding="utf-8-sig")


def test_an_export_without_registration_data_has_no_extra_sheet(tmp_path, patch_config):
    """沒用到開放資料的檔案不該多一張空的來源工作表。"""
    import openpyxl

    from exporter.service import get_exporter

    target = get_exporter("excel").export([_view(email="a@b.com")], tmp_path / "out.xlsx")

    assert "資料來源" not in openpyxl.load_workbook(target).sheetnames


def test_a_csv_with_the_attribution_is_still_a_valid_table(tmp_path, patch_config):
    """來源那一行補成合法的一列，不是 ``#`` 開頭的註解。CSV 沒有註解語法，
    寫成註解會讓嚴格一點的讀取器在最後一行報錯——而使用者匯出的檔案本來
    就是要拿去給別的程式讀的。"""
    import pandas as pd

    from exporter.service import get_exporter

    rows = [_view(id=1, registration_status="核准設立"), _view(id=2)]
    target = get_exporter("csv").export(rows, tmp_path / "out.csv")

    parsed = pd.read_csv(target)

    assert len(parsed) == len(rows) + 1  # 資料列 + 來源那一列
    assert OPEN_DATA_ATTRIBUTION in str(parsed.iloc[-1, 0])
