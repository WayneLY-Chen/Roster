"""Excel export with a formatted, filterable sheet."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.schemas import CompanyView
from exporter.base import BaseExporter, registry_attribution

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_MAX_COLUMN_WIDTH = 55
_MIN_COLUMN_WIDTH = 10


class ExcelExporter(BaseExporter):
    """Writes an .xlsx workbook with frozen headers and auto-filter."""

    extension = "xlsx"
    label = "Excel (.xlsx)"

    def _write(self, frame: pd.DataFrame, target: Path, rows: list[CompanyView]) -> None:
        sheet_name = self.config.exporter.excel_sheet_name[:31] or "Companies"

        with pd.ExcelWriter(target, engine="openpyxl") as writer:
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
            self._style(writer.sheets[sheet_name], frame)

            attribution = registry_attribution(rows)
            if attribution:
                # 開放資料的顯名標示（見 core.legal）。放在自己的工作表而不是
                # 資料表最下面補一列：資料表有自動篩選與凍結窗格，多出來的那
                # 一列會被算進篩選範圍，排序一下就跑到中間去了。
                pd.DataFrame({"": [attribution]}).to_excel(
                    writer, sheet_name="資料來源", index=False
                )
                writer.sheets["資料來源"].column_dimensions["A"].width = _MAX_COLUMN_WIDTH

    @staticmethod
    def _style(sheet, frame: pd.DataFrame) -> None:
        for cell in sheet[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        sheet.freeze_panes = "A2"
        if len(frame.columns):
            last_column = get_column_letter(len(frame.columns))
            sheet.auto_filter.ref = f"A1:{last_column}{len(frame) + 1}"

        for index, column in enumerate(frame.columns, start=1):
            # CJK glyphs are roughly two Latin characters wide in Excel's grid.
            widths = [_display_width(str(column))]
            widths += [_display_width(str(v)) for v in frame[column].head(200)]
            width = min(max(max(widths) + 2, _MIN_COLUMN_WIDTH), _MAX_COLUMN_WIDTH)
            sheet.column_dimensions[get_column_letter(index)].width = width


def _display_width(text: str) -> int:
    return sum(2 if ord(char) > 0x2E80 else 1 for char in text)
